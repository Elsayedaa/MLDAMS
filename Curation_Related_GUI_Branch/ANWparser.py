#Updated as of 5/12/2021 for GUI compatability

import openpyxl
import requests
from io import BytesIO
import pandas as pd
import numpy as np
import datetime
import re
import numpy as np


class anw:
    def __init__(self):
        
        #Active Neuron Worksheet OneDrive link:
        self.link = 'https://hhmionline-my.sharepoint.com/:x:/g/personal/arshadic_hhmi_org/EVEGrY83BCNHuVFt8c2awiwB3pzqdR3UhZhLV0Z8--U_yA?e=baTlLh'

        self.anw = 0
        while self.anw == 0:
            print("Loading active neuron worksheet...")
            try:
                #retrieving excel file binary from the onedrive link
                link = self.link.replace(self.link[self.link.index('?')+1:], 'download=1')
                dl = requests.get(link)
                excelfile = BytesIO(dl.content)

                #active neuron worksheet is read in data_only mode to read cell data instead of functions
                #it is also read in read_only mode to optimize loading time
                self.anw = openpyxl.load_workbook(excelfile, data_only = True, read_only = True)

                sheetarray = np.array(self.anw.sheetnames)
                rmatch_bools = np.vectorize(lambda x: bool(re.search(r"\d{4}-\d{2}-\d{2}$",x)))
                rtrue_index = np.where(rmatch_bools(sheetarray) == True)
                self.sheets = list(sheetarray[rtrue_index])
                print("Done!")

            #new link can be pasted into the cl terminal if the original is broken
            except Exception as e:
                if 'File is not a zip file' in e.args:
                    print("Active Neuron Worksheet link is not giving a response, link may be broken.\nReplace the 'link' variable in ANWparser.py with the new link.")
                    print("\nOr...")
                    self.link = input("Enter the new Active Neuron Worksheet link here (Note, this is not a permanent fix.):")
                    print()

    def set_activesheet(self, *sample):            
        #running a loop to initialize the sample worksheet as a class variable
        active_sheet = 0
        if sample == ():
            while active_sheet == 0:
                try:
                    self.sample = input("Enter the sample you would like to parse.")
                    active_sheet = self.anw[self.sample]
                    self.ws = pd.DataFrame(active_sheet.values)
                    print(f"You are now accessing sample: {self.sample}")
                except KeyError:
                    print("That sample does not exist in the Active Neuron Worksheet. Please try again.")          
        else:
            try:
                self.sample = sample[0]
                active_sheet = self.anw[self.sample]
                self.ws = pd.DataFrame(active_sheet.values)
                print(f"You are now accessing sample: {self.sample}")  
            except KeyError:
                print("That sample does not exist in the Active Neuron Worksheet. Please try again.")
                return

        #anw must be close because read_only mode uses lazy loading
        #self.anw.close() #closing is handled at the end of the GUI script     
        #editing the dataframe for ease of parsing
        self.ws.columns = self.ws.iloc[0].values
        self.ws.drop(self.ws.index[0], inplace = True)
        self.ws.set_index('Neuron Name', inplace  = True)
    
        #merged cells get unmerged in pandas
        #so here, I rename the rows which contain second passes
        #the index for each second pass is the neuron tag of the first pass + "x"
       
        fullvals = np.where(self.ws.index.values != None)[0]
        secvals = fullvals + 1
        self.ws.index.values[secvals] = self.ws.index.values[fullvals] + 'x'
        
        #intializing the count variables by running 'List' methods
        self.secpassneededList = self.secpassneeded_make_list()
        self.secpassdoneList = self.secpassdone_make_list()
        self.needsdendritesList = self.needsdendrites_make_list()
        self.needssplitList = self.needssplit_make_list()
        self.needsconsensusList = self.needsconsensus_make_list()
        self.consensuscompleteList = self.consensuscomplete_make_list()
        
        self.secpassneededCount = len(self.secpassneededList)
        self.secpassdoneCount = len(self.secpassdoneList)
        self.needsdendritesCount = len(self.needsdendritesList)
        self.needsplitCount = len(self.needssplitList)
        self.needsconsensusCount = len(self.needsconsensusList)
        self.consensuscompleteCount = len(self.consensuscompleteList)
    
    #This method creates a list of all neurons that need a second pass    
    def secpassneeded_make_list(self):
        namelist = []
        #The function first finds all neurons marked as completed
        c = self.ws.loc[(self.ws["Annotator Progress"]=="Completed")]
        #For all neurons marked as completed, the function checks the second pass (neuron tag + "x") associated with the neuron
        for i,n in enumerate(c.index):
            #if the second has an annotator progress of "In Progress" or "Waiting", it is added to the return list
            if "x" not in n and (self.ws["Annotator Progress"][n+"x"]=="In Progress" or self.ws["Annotator Progress"][n+"x"]=="Waiting"):
                namelist.append(n)
        return namelist
    
    #This method echoes a report of all neurons that need a second pass   
    def secpassneeded(self):
        print(f"2nd passes needed = {self.secpassneededCount}:")
        for n in self.secpassneededList:
            print(n)
        print()        
        
    #This method creates a list of all neurons which have a completed second pass    
    def secpassdone_make_list(self):
        namelist = []
        #The function first finds all neurons marked as completed
        c = self.ws.loc[(self.ws["Annotator Progress"]=="Completed")]
        #For all neurons marked as completed, the function checks if it has a completed second pass (neuron tag + "x")
        for n in c.index:
            #if the completed second pass is there, the neuron is added to the list
            if "x" in n and self.ws["Annotator Progress"][n.strip("x")] == "Completed":
                namelist.append(n.strip("x"))
        return namelist
    
    #This method echoes a report of all neurons that have a completed second pass  
    def secpassdone(self):
        print(f"2nd passes done = {self.secpassdoneCount}:")
        for n in self.secpassdoneList:
            print(n)
        print()
        
    #This method creates a list of dendrites that need to be completed
    def needsdendrites_make_list(self):
        #The method first creates a list of all dendrites marked as "Hold" or "Waiting"
        opendendrites = [n for n in self.ws.loc[(self.ws["Dendrites"]=="HOLD") | (self.ws["Dendrites"]=="Waiting") | (self.ws["Dendrites"]=="In Progress")].index if str(n).isnumeric()==False]
        namelist = []
        for n in opendendrites:
            #All dendrites which are also in the second pass done list are marked as needed with a finished second pass
            if n in self.secpassdoneList and self.ws["Dendrites"][n] == "In Progress":
                namelist.append(f"{n}: 2nd Pass: Done; dendrites in progress.")
            elif n in self.secpassdoneList:
                namelist.append(f"{n}: 2nd Pass: Done")
            #All dendrites which are also in the second pass needed list are marked as needed with a pending second pass
            if n in self.secpassneededList and self.ws["Dendrites"][n] == "In Progress":
                namelist.append(f"{n}: 2nd Pass: Pending; dendrites in progress.")
            elif n in self.secpassneededList:
                namelist.append(f"{n}: 2nd Pass: Pending")
        return namelist
    
    #This method echoes a report of all neurons that have dendrites pending completion  
    def needsdendrites(self):
        print(f"Dendrites needed = {self.needsdendritesCount}:")
        for n in self.needsdendritesList:
            print(n)
        print()
        
    #This method creates a list of all neurons that need to be split
    def needssplit_make_list(self):
        namelist = []
        #the method first parses through the second pass done list
        for n in self.secpassdoneList:
            #if the Consensus date value for any of these neurons are "none" or "split in progress" they are added to the return list
            if self.ws["Consensus Date"][n] == None or self.ws["Consensus Date"][n] == "Split In Progress":
                namelist.append(n)
        return namelist
    
    #This method echoes a report of all neurons that need to be split
    #Keep in mind that neurons will only show up here if they have a completed second pass
    def needssplit(self):
        print(f"Need to be split = {self.needsplitCount}:")
        for n in self.needssplitList:
            print(n)
        print()
        
    #This method creates a list of all neurons that need a consensus
    def needsconsensus_make_list(self):
        namelist = []
        #The method first creates a groupby object of neuron names and consensus date values
        consensus_dates = self.ws.groupby(self.ws.index)["Consensus Date"].value_counts()
        #Then the name/date pairs are iterated
        for sample, con in consensus_dates.index:
            #names that have a date value = "Split" are added to the return list
            if con == "Split":
                namelist.append(sample)
        return namelist
    
    #This method echoes a report of all neurons that need a consensus
    #Keep in mind that neurons will only show up here if they have been split
    def needsconsensus(self):
        print(f"Needs consensus = {self.needsconsensusCount}:")
        for n in self.needsconsensusList:
            print(n)
        print()
        
    #This method creates a list of all neurons that have a completed consensus
    def consensuscomplete_make_list(self):
        namelist = []
        #The method first creates a groupby object of neuron names and consensus date values
        consensus_dates = self.ws.groupby(self.ws.index)["Consensus Date"].value_counts()
        #Then the name/date pairs are iterated
        for neuron, con in consensus_dates.index:
            #names that have a date value which is a datetime object and are in the second pass done list are added to the return list
            if isinstance(con, datetime.date) and neuron in self.secpassdoneList: 
                namelist.append(neuron)
        return namelist
    
    #This method echoes a report of all neurons that have a completed consensus
    def consensuscomplete(self):
        print(f"Consensus complete = {self.consensuscompleteCount}:")
        for n in self.consensuscompleteList:
            print(n)
        print()

    #This method creates a list of coordinates for all neurons that have a completed consensus
    def coords_make_list(self):
        #The method consists of a list comprehention which removes brackets from the coordinates
        #It takes the series created by accessing the Neuron location column of the dataframe with a list
        #of neurons which have a completed consensus
        namelist = [c.strip(" ").strip("[]") for c in list(self.ws["Neuron Location (µm)"][self.consensuscompleteList].values)]
        return namelist
    
    #This method echoes a report of coordinates for all neurons that have a completed consensus
    def returncoords(self):
        print(f"Coordinates:")
        for n in self.coords_make_list():
            print(n)
        print()
        
    #This method echoes the percent of sample completed 
    #calculated as a proportion of neurons with a consensus over neurons with a second pass started or finished
    def percentcomplete(self):
        total = int(self.secpassneededCount + self.secpassdoneCount)
        try:
            print(f"Percent complete: {(int(self.consensuscompleteCount)/total)*100} %")
        except ZeroDivisionError:
            print(f"Cannot calculate completeness before the first neuron has been \nstarted.")
    #This method returns a summary report including all methods except returncoords
    def all(self):
        self.secpassdone()
        self.secpassneeded()
        self.needssplit()
        self.needsdendrites()
        self.needsconsensus()
        self.consensuscomplete()
        self.percentcomplete()